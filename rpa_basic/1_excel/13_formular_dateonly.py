from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 데이터를 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]
for s in scores:
    ws.append(s)

# 1. 퀴즈2 점수를 10으로 수정
for index, cell in enumerate(ws["D"]):
    if index == 0: # 제목인 경우
        continue
    cell.value = 10

# 2. H열에 총점(sum) 이용, I 열에 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value = "=sum(B{}:G{})".format(idx, idx)

    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade # I 열에 성적 정보

wb.save("scores.xlsx")

# wb = load_workbook("sample_formular.xlsx")
# ws = wb.active


# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

#수식이 아닌 실제 데이터를 가지고 옴
# # evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# wb = load_workbook("sample_formular.xlsx", data_only=True)
# ws = wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)

# wb.save("sample_formular.xlsx")
